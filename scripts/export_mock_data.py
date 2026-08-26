import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.data.catalog import ASSETS, CONTROLS, DECOYS, PATH_DEFS, THREATS, VULNERABILITIES


DATASETS = {
    "Assets": ASSETS,
    "Vulnerabilities": VULNERABILITIES,
    "Threats": THREATS,
    "Decoys": DECOYS,
    "Controls": CONTROLS,
    "Attack Paths": PATH_DEFS,
}


def export_workbook(output_path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="0F766E")

    for sheet_name, records in DATASETS.items():
        sheet = workbook.create_sheet(sheet_name)
        keys = list(dict.fromkeys(key for record in records for key in record))
        sheet.append(keys)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for record in records:
            sheet.append([
                json.dumps(record.get(key), ensure_ascii=True)
                if isinstance(record.get(key), (list, dict))
                else record.get(key)
                for key in keys
            ])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 48)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width

    workbook.save(output_path)


if __name__ == "__main__":
    export_workbook(Path(__file__).resolve().parents[1] / "mock-data.xlsx")